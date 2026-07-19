# -*- coding: utf-8 -*-
"""
jobkorea_bot_main.py
잡코리아 수집결과에서 '타겟 기업' 공고만 골라내는 스크립트

흐름
  1) jobkorea_list.csv       : 잡코리아 봇이 수집한 전체 공고 (입력)
  2) target_companies.xlsx   : 타겟 기업 목록 (Sheet1 A열 '기업명'만 사용)
  3) 두 파일의 기업명을 정규화 매칭 -> 타겟에 해당하는 공고만
  4) jobkorea_main.csv       : 매칭된 공고만 저장 (출력)

매칭 규칙 (C안 절충)
  - 양쪽 회사명을 정규화: ㈜·(주)·주식회사·(유) 등 법인표기/공백/특수문자 제거, 영문 소문자화
  - 정규화된 '타겟명' 길이 기준:
      * 2글자 이하(GC, AK 등)  -> 완전일치만 인정 (오탐 방지)
      * 3글자 이상             -> 완전일치 또는 '타겟명 + 접미사'(타겟명으로 시작)면 인정
  - 예) 동국제약 ↔ 동국제약㈜ (O) / GS리테일 ↔ GS리테일GS그룹 (O) / GC ↔ GC이엔지 (X)

필요 라이브러리:
    pip install openpyxl

참고
  - 같은 폴더에 세 파일이 있다고 가정합니다.
  - target A열만 쓰므로 영문 타겟(Google 등)은 국문 법인명 공고와 매칭되지 않습니다.
"""

import os
import csv
import re

import openpyxl

# ===========================================================================
# 설정
# ===========================================================================
INPUT_CSV = "jobkorea_list.csv"          # 잡코리아 전체 수집 결과 (입력)
TARGET_XLSX = "target_companies.xlsx"    # 타겟 기업 목록
OUTPUT_CSV = "jobkorea_main.csv"         # 매칭 결과 (출력)

TARGET_SHEET = "Sheet1"                   # 타겟 시트명 (없으면 첫 시트 사용)
TARGET_COMPANY_HEADER = "기업명"          # 타겟에서 사용할 A열 헤더
CSV_COMPANY_COL = "기업명"                # 잡코리아 CSV의 회사명 컬럼

ADD_MATCH_COLUMN = True                   # 끝에 '매칭타겟명' 컬럼 추가할지
MATCH_COLUMN_NAME = "매칭타겟명"

# 정규화 시 제거할 법인/조직 표기
LEGAL_TOKENS = [
    "주식회사", "유한회사", "유한책임회사", "재단법인", "사단법인",
    "의료법인", "학교법인", "(주)", "（주）", "㈜", "(유)", "(재)",
    "(사)", "(합)", "(유한)", "(주식회사)",
]


# ===========================================================================
# 회사명 정규화
# ===========================================================================
def normalize(name):
    """법인표기/공백/특수문자 제거 + 영문 소문자화 -> 비교용 문자열."""
    if not name:
        return ""
    s = str(name).lower()
    for tok in LEGAL_TOKENS:
        s = s.replace(tok.lower(), "")
    # 한글/영문/숫자만 남김 (공백·괄호·점·중점 등 제거)
    s = re.sub(r"[^0-9a-z가-힣]", "", s)
    return s


# ===========================================================================
# 타겟 목록 로드 (A열 기업명만)
# ===========================================================================
def load_targets(path):
    """반환: {정규화명: 원본표시명}  (중복은 첫 등장 우선)"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[TARGET_SHEET] if TARGET_SHEET in wb.sheetnames else wb.worksheets[0]

    target_map = {}
    header_seen = False
    for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        raw = row[0]
        if raw is None or str(raw).strip() == "":
            continue
        name = str(raw).strip()
        # 헤더 행 스킵
        if not header_seen and name == TARGET_COMPANY_HEADER:
            header_seen = True
            continue
        norm = normalize(name)
        if norm and norm not in target_map:
            target_map[norm] = name
    wb.close()
    print(f"[타겟] {len(target_map)}개 기업 로드 (중복 제거 후)")
    return target_map


# ===========================================================================
# 매칭
# ===========================================================================
def match_company(norm_company, target_map):
    """
    C안 절충 매칭. 매칭되면 (원본 타겟명) 반환, 아니면 None.
      - 완전일치: 모든 길이의 타겟에 대해 인정
      - 접두일치: 3글자 이상 타겟에 대해서만 인정 (norm_company 가 타겟명 + 접미사)
    """
    if not norm_company:
        return None
    # 1) 완전일치 (길이 무관)
    if norm_company in target_map:
        return target_map[norm_company]
    # 2) 접두일치 (3글자 이상 타겟만) -> company 의 3글자 이상 접두부가 타겟과 일치
    for L in range(3, len(norm_company)):
        prefix = norm_company[:L]
        if prefix in target_map:
            return target_map[prefix]
    return None


# ===========================================================================
# 메인
# ===========================================================================
def main():
    if not os.path.exists(INPUT_CSV):
        print(f"[오류] 입력 파일 없음: {INPUT_CSV}")
        return
    if not os.path.exists(TARGET_XLSX):
        print(f"[오류] 타겟 파일 없음: {TARGET_XLSX}")
        return

    target_map = load_targets(TARGET_XLSX)

    with open(INPUT_CSV, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        fieldnames = list(reader.fieldnames or [])
        if CSV_COMPANY_COL not in fieldnames:
            print(f"[오류] CSV에 '{CSV_COMPANY_COL}' 컬럼이 없습니다. 컬럼: {fieldnames}")
            return
        rows = list(reader)

    print(f"[입력] {INPUT_CSV}: {len(rows)}건")

    out_fields = fieldnames + ([MATCH_COLUMN_NAME] if ADD_MATCH_COLUMN else [])
    matched = []
    for row in rows:
        company = (row.get(CSV_COMPANY_COL) or "").strip()
        if not company:
            continue
        hit = match_company(normalize(company), target_map)
        if hit:
            if ADD_MATCH_COLUMN:
                row[MATCH_COLUMN_NAME] = hit
            matched.append(row)

    with open(OUTPUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=out_fields)
        writer.writeheader()
        for r in matched:
            writer.writerow({k: r.get(k, "") for k in out_fields})

    print(f"[출력] {OUTPUT_CSV}: 매칭 {len(matched)}건 저장")
    # 매칭 결과 미리보기 (상위 10건)
    for r in matched[:10]:
        tag = f"  <- {r.get(MATCH_COLUMN_NAME)}" if ADD_MATCH_COLUMN else ""
        print(f"  · {r.get(CSV_COMPANY_COL)} | {r.get('공고제목', '')}{tag}")


if __name__ == "__main__":
    main()